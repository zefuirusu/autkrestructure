#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
import os
from threading import Thread
from xlrd import open_workbook
from openpyxl import load_workbook
from numpy import array,zeros
from pandas import read_excel,DataFrame

from autk.gentk.funcs import regex_filter,start_thread_list
from autk.mapper.base import XlMap

class XlBook:
    '''
    Basic Structure of XlBook on default:
    `xls` starts from 0, while `xlsx` starts from 1;
        file_path,
        file_name,
        suffix:xls/xlsx/xlsm,
        shtli,
        data,
    '''
    def __init__(self,file_path):
        self.file_path=file_path
        self.file_name=''
        self.pure_file_name=''
        self.suffix=''
        self.__parse_file_type()
        pass
    def __parse_file_type(self):
        self.file_name=self.file_path.split(os.sep)[-1]
        self.pure_file_name=re.sub(
            re.compile(r'\.xls[xm]?$'),
            '',
            self.file_name
        )
        self.suffix=re.sub(
            re.compile(r'^.*\.'),
            '',
            self.file_name
        )
        pass
    @property
    def shtli(self):
        if self.suffix=='xls':
            return [sht.name for sht in open_workbook(self.file_path).sheets()]
        elif self.suffix=='xlsx':
            return load_workbook(self.file_path).sheetnames
        elif self.suffix=='xlsm':
            return load_workbook(self.file_path,keep_vba=True).sheetnames
        pass
    @property
    def shape(self):
        '''
        sheet_name | shape
        -----------|-------
        sheet_1    | (n_1,m_1)
        sheet_2    | (n_2,m_2)
        .......
        '''
        if self.suffix=='xls':
            shape=array(
                [
                    [open_workbook(self.file_path).sheet_by_name(sht).nrows,
                     open_workbook(self.file_path).sheet_by_name(sht).ncols]
                    for sht in self.shtli
                ]
            )
            pass
        elif self.suffix=='xlsx' or 'xlsm':
            shape=array(
                [
                    [load_workbook(self.file_path)[sht].max_row,
                     load_workbook(self.file_path)[sht].max_column]
                    for sht in self.shtli
                ]
            )
            pass
        return shape
    @property
    def shape_df(self):
        return DataFrame(
            data=self.shape,
            columns=['rows','cols'],
            index=self.shtli
        )
    def test_map(self,xlmap,common_title=0):
        '''
        To test each sheet to check if they are fit to the input `xlmap`;
        If error appears, there may be some hidden sheets, within that Excel,
        whose columns conflict with that of the unhidden ones.
        '''
        print('start testing map......\n','If error appears, there may be some hidden sheets, within that Excel,whose columns conflict with that of the unhidden ones.')
        resu_df=DataFrame([],index=self.shape_df.index,columns=xlmap.columns)
        map_cols=xlmap.columns
        map_dict=xlmap.show
        for sht in resu_df.index:
            max_cols=self.shape_df.at[sht,'cols']
            sht_cols=self.get_row(sht,common_title+1)
            for col in resu_df.columns:
                col_index=map_dict[col]
                if col_index is not None:
                    resu_df.at[sht,col]=sht_cols[col_index]
                continue
            continue
        return resu_df
    def save_bk(self,bk):
        if self.suffix=='xlsx' or 'xlsm':
            bk.save(self.file_path)
        else:
            bk.save(self.file_path)
        pass
    def get_bk(self,write=False):
        if self.suffix=='xlsx':
            if write==True:
                bk=load_workbook(self.file_path,keep_links=True,rich_text=False)
            else:
                bk=load_workbook(self.file_path,data_only=True,keep_links=True,rich_text=False)
        elif self.suffix=='xlsm':
            if write==True:
                bk=load_workbook(self.file_path,keep_vba=True,keep_links=True,rich_text=False)
            else:
                bk=load_workbook(self.file_path,keep_vba=True,data_only=True,keep_links=True,rich_text=False)
        elif self.suffix=='xls':
            bk=open_workbook(self.file_path)
        else:
            bk=open_workbook(self.file_path)
        return bk
    def find_sheet(self,regex_str):
        possible_names=regex_filter(
            regex_str,
            self.shtli,
            match_mode=False
        )
        return possible_names
    def get_sht(self,sheet_name):
        if self.suffix=='xlsx' or 'xlsm':
            sht=self.get_bk()[sheet_name]
        elif self.suffix=='xls':
            sht=self.get_bk().sheet_by_name(sheet_name)
        else:
            sht=None
        return sht
    def search_in_sheet(self,cell_content,sheet_name):
        def __sheet_search_xlsx(cell_content,shtna):
            cell_content=re.compile(cell_content)
            sht=self.get_sht(shtna)
            sht_resu=[]
            for r in range(sht.max_row):
                for c in range(sht.max_column):
                    mt=re.search(cell_content,str(sht.cell(r+1,c+1).value))
                    if mt is not None:
                        sht_resu.append((r+1,c+1))
                    else:
                        pass
            return sht_resu
        def __sheet_search_xls(cell_content,shtna):
            cell_content=re.compile(cell_content)
            sht=self.get_sht(shtna)
            sht_resu=[]
            for r in range(sht.nrows):
                for c in range(sht.ncols):
                    mt=re.search(cell_content,str(sht.cell(r,c).value))
                    if mt is not None:
                        sht_resu.append((r,c))
                    else:
                        pass
            return sht_resu
        if self.suffix=='xls':
            return __sheet_search_xls(cell_content,sheet_name)
        else:
            return __sheet_search_xlsx(cell_content,sheet_name)
    def search(self,cell_content):
        '''
        Search cells by its content and return index of the result.
        parameters:
            cell_content:regex string;
        returns:
            `xls` starts from 0, while `xlsx` starts from 1;
            {
                "sheet_name A":[(row1,column1),(row2,column2),...],
                "sheet_name B":[(row3,column3),...],
                ...
            }
        '''
        resu={}
        def __sheet_search(cell_content,shtna):
            sht_resu=self.search_in_sheet(cell_content,shtna)
            if len(sht_resu)>0:
                resu.update({shtna:sht_resu})
            else:
                pass
        thli=[]
        for shtna in self.shtli:
            thli.append(
                Thread(
                    target=__sheet_search,
                    args=(cell_content,shtna)
                )
            )
            continue
        start_thread_list(thli)
        return resu
    def get_value(self,sheet_name,cell_index):
        '''
        `xls` starts from 0, while `xlsx` starts from 1;
        '''
        if self.suffix=='xlsx' or 'xlsm':
            value=self.get_sht(sheet_name).cell(
                cell_index[0],
                cell_index[1]
            ).value
        elif self.suffix=='xls':
            value=self.get_sht(sheet_name).cell(
                cell_index[0],
                cell_index[1]
            ).value
            pass
        else:
            value=0
        return value
    def get_matrix(
            self,
            sheet_name,
            start_cell_index,
            n_rows_range,
            n_cols_range,
            type_df=False,
            has_title=False
    ):
        '''
        start_cell_index is a tuple like 'R1C1' ref-style in Excel: (row,column);
        For xlrd.open_workbook().sheet_by_name(), index starts from 0;
        While for openpyxl.load_workbook().get_sheet_by_name(), index starts from 1;
        That's all right, just start from 1 when passing argument 'start_cell_index' as tuple like (n,m).
        For numbers, different file type results in different data type:
            xls:str(float)
            xlsx/xlsm:str(int)
        '''
        sht=self.get_sht(sheet_name)
        if self.suffix=='xls':
            #  sht=open_workbook(self.file_path).sheet_by_name(sheet_name)
            matrix=array(
                [
                    sht.row_values(
                        row,
                        start_cell_index[1]-1,
                        start_cell_index[1]-1+n_cols_range
                    ) for row in range(
                        start_cell_index[0]-1,
                        start_cell_index[0]-1+n_rows_range
                    )
                ]
            )
        elif self.suffix=='xlsx':
            #  sht=load_workbook(self.file_path).get_sheet_by_name(sheet_name) # same as:
            matrix=array(
                list(
                    sht.iter_rows(
                        min_row=start_cell_index[0],
                        max_row=start_cell_index[0]+n_rows_range-1,
                        min_col=start_cell_index[1],
                        max_col=start_cell_index[1]+n_cols_range-1,
                        values_only=True
                    )
                )
            )
        elif self.suffix=='xlsm':
            #  sht=load_workbook(self.file_path).get_sheet_by_name(sheet_name) # same as:
            matrix=array(
                list(
                    sht.iter_rows(
                        min_row=start_cell_index[0],
                        max_row=start_cell_index[0]+n_rows_range-1,
                        min_col=start_cell_index[1],
                        max_col=start_cell_index[1]+n_cols_range-1,
                        values_only=True
                    )
                )
            )
        else:
            matrix=zeros((n_rows_range,n_cols_range))
        if type_df==True:
            if has_title==False:
                matrix=DataFrame(
                    data=matrix
                )
            else:
                matrix=DataFrame(
                    data=matrix[1:],
                    columns=matrix[0]
                )
        else:
            pass
        return matrix
    def select_matrix(
        self,
        sheet_name,
        from_cell_index,
        to_cell_index,
        type_df=False,
        has_title=False
    ):
        '''
        from_cell_index and to_cell_index are tuples like 'R1C1' ref-style in Excel: (row,column);
        '''
        return self.get_matrix(
            sheet_name,
            from_cell_index,
            to_cell_index[0]-from_cell_index[0]+1,
            to_cell_index[1]-from_cell_index[1]+1,
            type_df=type_df,
            has_title=has_title
        )
    def select_all(self,sheet_name,type_df=False):
        return self.select_matrix(
            sheet_name,
            (1,1),
            (self.shape_df.at[sheet_name,'rows'],self.shape_df.at[sheet_name,'cols']),
            type_df=type_df,
            has_title=False if type_df==False else True
        )
    def get_row(self,sheet_name,row):
        max_col=self.shape_df.at[sheet_name,'cols']
        return list(
            self.select_matrix(
                sheet_name,
                (row,1),
                (row,max_col),
                type_df=False,
                has_title=False
            )[0]
        )
    def get_col(self,sheet_name,col):
        max_row=self.shape_df.at[sheet_name,'rows']
        return list(
            self.select_matrix(
                sheet_name,
                (1,col),
                (max_row,col),
                type_df=False,
                has_title=False
            ).T[0]
        )
    def __xls_fill(self,sheet_name,cell_index,value,save=False):
        import xlutils
        #  b=open_workbook(self.file_path)
        b=self.get_bk(write=True)
        b=xlutils.copy(b)
        #  b.sheet_by_name(sheet_name).cell(cell_index[0],cell_index[1]).value=value
        b.write(cell_index[0],cell_index[1],value,save=False)
        if save==True:
            self.save_bk(b)
            #  b.save(self.file_path)
        pass
    def __xlsx_fill(self,sheet_name,cell_index,value,save=False):
        #  b=load_workbook(self.file_path)
        b=self.get_bk(write=True)
        b[sheet_name].cell(row=cell_index[0],column=cell_index[1]).value=value
        if save==True:
            self.save_bk(b)
            #  b.save(self.file_path)
        pass
    def __xlsm_fill(self,sheet_name,cell_index,value,save=False):
        #  b=load_workbook(self.file_path,keep_vba=True)
        b=self.get_bk(write=True)
        b[sheet_name].cell(row=cell_index[0],column=cell_index[1]).value=value
        if save==True:
            self.save_bk(b)
            #  b.save(self.file_path)
        pass
    def get_fill_func(self,save=False):
        '''
        save=False by default;
        '''
        if self.suffix=='xlsx':
            def fill_func(sheet_name,row_num,col_num,value):
                self.__xlsx_fill(sheet_name,(row_num,col_num),value,save=save)
                pass
            pass
        elif self.suffix=='xlsm':
            def fill_func(sheet_name,row_num,col_num,value):
                self.__xlsm_fill(sheet_name,(row_num,col_num),value,save=save)
                pass
            pass
        elif self.suffix=='xls':
            def fill_func(sheet_name,row_num,col_num,value):
                self.__xls_fill(sheet_name,(row_num,col_num),value,save=save)
                pass
            pass
        else:
            def fill_func(sheet_name,row_num,col_num,value):
                self.__xlsx_fill(sheet_name,(row_num,col_num),value,save=save)
                pass
        return fill_func
    def fill_value(self,sheet_name,cell_index,value):
        '''
        fill a single value and then save;
        '''
        self.get_fill_func(save=True)(sheet_name,cell_index[0],cell_index[1],value)
        pass
    def fill_bydf(self,matrix):
        '''
        matrix could be: DataFrame,array, or 2-dimension list;
        _________________________________________
        |sheet_name|row_index|column_index|value|
        -----------------------------------------
        |__________|_________|____________|_____|
        '''
        #  from threading import Thread
        #  thread_list=[]
        if isinstance(matrix,DataFrame):
            cols=matrix.columns
            for row in matrix.iterrows():
                row_data=row[1]
                sht=row_data[cols[0]]
                r_index=row_data[cols[1]]
                c_index=row_data[cols[2]]
                v=row_data[cols[3]]
                self.get_fill_func(save=True)(sht,r_index,c_index,v)
                #  thread_list.append(
                    #  Thread(
                        #  target=self.get_fill_func(save=True),
                        #  args=(sht,r_index,c_index,v)
                    #  )
                #  )
            pass
        elif isinstance(matrix,list) or isinstance(matrix,array):
            for row in matrix:
                sht=row[0]
                r_index=row[1]
                c_index=row[2]
                v=row[3]
                self.get_fill_func(save=True)(sht,r_index,c_index,v)
                #  thread_list.append(
                    #  Thread(
                        #  target=self.get_fill_func(save=True),
                        #  args=(sht,r_index,c_index,v)
                    #  )
                #  )
            pass
        else:
            #  thread_list.append(
                #  Thread(
                    #  target=print,
                    #  args=('check your arguments:matrix',),
                #  )
            #  )
            pass
        #  start_thread_list(thread_list)
        pass
    def paste_matrix(self,matrix,start_index,sheet_name):
        '''
        matrix:
            must be nd.array or pandas.DataFrame;
        start_index:
            from which cell of the `sheet_name` to start;
        '''
        r=start_index[0]
        c=start_index[1]
        from numpy import ndarray,nditer
        if isinstance(matrix,ndarray):
            for row in matrix:
                for value in row:
                    self.fill_value(sheet_name,(r,c),value)
                    c+=1
                    continue
                c=start_index[1]
                r+=1
                continue
            pass
        elif isinstance(matrix,DataFrame):
            matrix=matrix.values
            self.paste_matrix(matrix,start_index,sheet_name)
            pass
        else:
            print('You may check argument: ',matrix)
            matrix=array(matrix)
            self.paste_matrix(
                matrix,
                start_index,
                sheet_name
            )
            pass
        pass
    def paste_list(
        self,
        value_list,
        start_index,
        sheet_name,
        vertical=True
    ):
        value_list=array([value_list])
        if vertical==False:
            pass
        else:
            value_list=value_list.T
        self.paste_matrix(
            value_list,
            start_index,
            sheet_name
        )
        pass
    def insert_blank(
        self,
        idx,
        amount,
        sheet_name,
        col_ins=False
    ):
        '''
        Insert `amount` rows/columns before row_index/col_index:`idx`
        into `sheet_name`;
        '''
        bk=self.get_bk()
        if (
            self.suffix=='xlsx' or 
            self.suffix=='xlsm'
        ):
            sht=bk[sheet_name]
            if col_ins==True:
                sht.insert_cols(idx,amount=amount)
            else: # insert rows
                sht.insert_rows(idx,amount=amount)
            pass
        elif self.suffix=='xls':
            #  sht=bk.sheet_by_name(sheet_name)
            print(
                '[Error] Inserting to `xls` is not supported currently.'
            )
            pass
        self.save_bk(bk)
        pass
    def insert_matrix(
        self,
        matrix, # ndarray;
        cell_index:(int,int),
        sheet_name:str,
        col_ins:bool=False,
    ):
        '''
        `cell_index` starts from (1,1);
        '''
        from numpy import ndarray,nditer
        if isinstance(matrix,ndarray):
            pass
        elif isinstance(matrix,DataFrame):
            matrix=matrix.values
        elif isinstance(matrix,list) and matrix.ndim==2:
            matrix=array(matrix)
        else:
            pass
        if col_ins==True:
            idx=cell_index[1]
            amount=matrix.shape[1]
            pass
        else: # insert rows and paste;
            idx=cell_index[0]
            amount=matrix.shape[0]
            pass
        self.insert_blank(
            #  cell_index[0],
            idx,
            matrix.shape[0],
            sheet_name,
            col_ins=col_ins
        )
        self.paste_matrix(
            matrix,
            cell_index,
            sheet_name
        )
        pass
    def clear_sheet(self,sheet_name):
        '''
        Carefull! This method will clear all data of the target sheet!
        '''
        from numpy import full
        z=full(self.shape[0],'')
        self.paste_matrix(z,(1,1),sheet_name)
        pass
    def get_df(self,sheet_name,title=0):
        if self.suffix==r'xls':
            return read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=title,
                engine='xlrd'
            )
        elif self.suffix==r'xlsx':
            return read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=title,
                engine='openpyxl'
            )
        elif self.suffix==r'xlsm':
            return read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=titile,
                engine='openpyxl'
            )
        else:
            print(
                "[Error]:check the input file:",
                self.file_path
            )
            return DataFrame([])
    def get_mapdf(
        self,
        sheet_name:str,
        xlmap:XlMap,
        title=0
    ):
        from copy import deepcopy
        source_data=self.get_df(sheet_name,title=title)
        data=DataFrame(
            [],
            columns=xlmap.columns
        )
        print(
            '[{}] xlmap show:{}'.format(
                self.__class__.__name__,
                xlmap.show
            )
        )
        for col in xlmap.columns:
            col_index=xlmap.show[col]
            if isinstance(col_index,int):
                print(
                    '[{}] draw column `{}` at location `{}` from source data.'.format(
                        self.__class__.__name__,
                        col,
                        col_index
                    )
                )
                col_from_source=source_data.columns.to_numpy()[col_index]
                data[col]=deepcopy(
                    source_data[col_from_source]
                )
            elif isinstance(col_index,list):
                data[col]=0
                print(
                    '[{}] draw column `{}` at location `{}` from source data.'.format(
                        self.__class__.__name__,
                        col,
                        col_index
                    )
                )
                for sub_col_index in col_index:
                    sub_col_from_source=source_data.columns[sub_col_index]
                    data[col]=deepcopy(
                        data[col]+source_data[sub_col_from_source]
                    )
                    continue
            else:
                pass
            continue
        return data
    def to_sht(
        self,
        sheet_name:str,
        xlmap:XlMap=None,
        common_title=0,
    ):
        from autk.calculation.base.xlsht import XlSheet
        from autk.meta.pmeta import PathMeta
        xl=XlSheet(
            xlmap=xlmap,
            shmeta=PathMeta(
                self.file_path,
                shtli=[sheet_name],
                common_title=common_title,
                keep_additional=False,
            ),
        )
        return xl
    def to_mtb(
        self,
        common_title=0,
        auto_load=False
    ):
        '''
        Transform self into ImmortalTable.
        '''
        from autk.calculation.base.table import ImmortalTable
        xlmeta={}
        xlmeta.update(
            {self.file_path:[[sht,common_title] for sht in
                             self.shtli]}
        )
        return ImmortalTable(
            xlmap=None,
            xlmeta=xlmeta,
        )
        pass
    def to_mgl(
        self,
        common_title=0,
        xlmap=None,
        auto_load=False
    ):
        from autk.calculation.mortal.mortalgl import MGL
        xlmeta={}
        xlmeta.update(
            {self.file_path:[[sht,common_title] for sht in
                             self.shtli]}
        )
        return MGL(
            xlmap=None,
            xlmeta=xlmeta,
        )
    def to_chart(
        self,
        common_title=0,
        xlmap=None,
        auto_load=False
    ):
        from autk.calculation.mortal.mortalchart import MCH
        xlmeta={}
        xlmeta.update(
            {self.file_path:[[sht,common_title] for sht in 
                             self.shtli]}
        )
        return MCH(
            xlmap=None,
            xlmeta=xlmeta,
        )
        pass
    def to_inventory(
        self,
    ):
        pass
    pass
if __name__=='__main__':
    pass
