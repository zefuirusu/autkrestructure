#!/usr/bin/env python
# encoding = 'utf-8'
from autk.gentk.start import startprj,start_by_conf
from autk.meta.handf.pdfsplit import multi_split

def start_thread_list(thread_list):
    for t in thread_list:
        t.start()
    for t in thread_list:
        t.join()
    pass
def f2list(file_path):
    import re
    with open(file_path,mode='r',encoding='utf-8') as f:
        fli=f.readlines()
    return list(map(lambda x:re.sub(r'\s*$', '', x),fli))
def f2dict(file_path):
    import json
    with open(file_path,encoding='utf-8') as f:
        fjson=json.load(f)
    return fjson
def dir2json(base_dir,shtna='表页-1',title=3):
    import os
    file_li=os.listdir(base_dir)
    dict_meta={}
    for file_name in file_li:
        file_path=os.path.abspath(
            os.path.join(base_dir,file_name)
        )
        dict_meta.update(
            {file_path:[[shtna,title]]}
        )
        continue
    return dict_meta
    pass
def relative_path(target,reference):
    '''
    What is the relative path of 'target' file in sight of 'reference' file ?
    How to get 'target' file location from 'reference' file ?
    ----
    Return the relative path of 'target' towards 'reference'.
    Find the location relationship between two `files`.
    ----
    This function is being tested.
    '''
    import os
    from copy import deepcopy
    target=os.path.abspath(target)
    reference=os.path.abspath(reference)
    #  print('target:\n',target)
    #  print('ref:\n',reference)
    target_split=target.split(os.sep)
    ref_split=reference.split(os.sep)
    def remove_common(left_list,right_list):
        '''
        Assume that left_list is shorter than right_list;
        '''
        left_list_copy=deepcopy(left_list)
        right_list_copy=deepcopy(right_list)
        common_count=len(
            list(
                set(left_list_copy)&set(right_list_copy)
            )
        )
        for k in range(common_count):
            left_list_copy.pop(0)
            right_list_copy.pop(0)
            continue
        return [common_count,left_list_copy,right_list_copy]
    common_count=remove_common(ref_split,target_split)[0]
    ref_list_copy=remove_common(ref_split,target_split)[1]
    target_list_copy=remove_common(ref_split,target_split)[2]
    up_step=len(ref_split)-common_count-1
    if up_step==0:
        resu_path=[r'.']
    else:
        resu_path=[r'..']*(up_step)
    resu_path.extend(target_list_copy)
    resu_path=os.sep.join(resu_path)
    #  print('result:\n',resu_path)
    return resu_path
def regex_filter(regex_str,list_like,match_mode=False):
    '''
    list_like is a list-like iterable object, whose elements are 'str';
    '''
    import re
    regex_str=re.compile(regex_str)
    if match_mode == False:
        def regex_result(element):
            return (re.search(regex_str,element) is not None)
    else:
        def regex_result(element):
            return (re.match(regex_str,element) is not None)
    return list(filter(regex_result,list_like))
def regex_filter_dict(regex_str,dict_like,key_resu=False,parse_key=False,match_mode=False):
    import re
    regex_str=re.compile(regex_str)
    pass
def get_time_str(woc=False):
    '''
    parameters:
        woc: bool
        if True: '20210710-212154'
        if False: '2021_07_10-21_21_50'
    '''
    import time
    time_list=list(map(str,time.localtime()))
    if len(time_list[1])==1:
        time_list[1]='0'+time_list[1]
        pass
    else:
        pass
    if len(time_list[2])==1:
        time_list[2]='0'+time_list[2]
        pass
    else:
        pass
    if woc == False:
        timestr='-'.join(['_'.join(time_list[0:3]),'_'.join(time_list[3:6])])
    else:
        timestr='-'.join([''.join(time_list[0:3]),''.join(time_list[3:6])])
        # timestr=''.join(time_list[0:6])
    return timestr
def transType(element):
    '''
    transfer an float/integer object into string.
    '''
    import re
    element=str(element)
    if element=='nan':
        element='0'
    else:
        if re.match(r'.*\.0*$',element) is not None:
            element=re.sub(r'\.0*$','',element)
        else:
            pass
    return element
def wtlog(logline,logdir='./log_default.txt'):
    '''
    writing printing log file.
    logline is what you write and logdir is the log file location.
    '''
    with open(logdir,mode='a',encoding = 'utf-8') as f:
        f.write(logline)
        f.write('\n')
    return
def trans_to_xlsx(in_file_path,save_path):
    '''
    Transfer Excel '*.xls' to '*.xlsx'.
    module 'win32com' is required.
    Parameters
    ----------
    in_file_path : TYPE
        DESCRIPTION.
    save_path : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    # import os
    # import win32.client as winclient
    # e = winclient.Dispatch('Excel.application')
    # pro = e.Workbooks.Open(in_file_path)
    # # file_name=in_file_path.split(os.sep)[-1]
    # # new_file_name=re.sub(r'\.xls$',r'.xlsx',file_name)
    # pro.SaveAs(save_path, FileFormat=56)  # '56' representss 'xlsx'; '51' represents 'xls'.
    # pro.Close()
    # e.Application.Quit()
    pass
def trans_dir_xlsx(in_file_dir):
    from os import listdir
    from os.path import join as ojoin
    file_pathli=[]
    for f in listdir(in_file_dir):
        file_pathli.append(ojoin(in_file_dir,f))
    from threading import Thread
    thread_list=[]
    for f in file_pathli:
        t=Thread(target=trans_to_xlsx,args=(f,f))
    for t in thread_list:
        t.start()
    for t in thread_list:
        t.join()
    pass
def save_df(df,sheet_name=None,save_path='./',file_nickname='data'):
    if sheet_name is None:
        sheet_name=get_time_str(woc=True)
    from os.path import isfile,isdir
    from os.path import join as osjoin
    from pandas import ExcelWriter
    from openpyxl import Workbook,load_workbook
    def __save_to_path(save_path):
        wb=Workbook()
        wb.save(save_path)
        wb.close()
        df.to_excel(save_path,sheet_name=sheet_name)
        print('\n','saved to path:',save_path,'\n','sheet_name:',sheet_name,'\n')
    def __save_to_file(save_path):
        #  wb=load_workbook(save_path,read_only=False, keep_vba=True, data_only=False, keep_links=True)
        wter=ExcelWriter(
            save_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='new',
            engine_kwargs={
                #  "filename":save_path,
                "read_only":False,
                "keep_vba":True,
                "data_only":False,
                "keep_links":False
            }
        )
        df.to_excel(wter,sheet_name=sheet_name)
        wter.close()
        print('\n','saved to file:',save_path,'\n','sheet_name:',sheet_name,'\n')
    def __save_to_dir(save_path):
        file_name=''.join([file_nickname,get_time_str(woc=True),r'.xlsx'])
        save_path=osjoin(save_path,file_name)
        if isfile(save_path):
            __save_to_file(save_path)
        else:
            wb=Workbook()
            wb.save(save_path)
            wter=ExcelWriter(save_path)
            wter.book=wb
            df.to_excel(wter,sheet_name=sheet_name)
            wter.save()
        print('\n','saved to dir:',save_path,'\n','sheet_name:',sheet_name,'\n')
    if isfile(save_path):
        __save_to_file(save_path)
        pass
    elif isdir(save_path):
        __save_to_dir(save_path)
        pass
    elif isinstance(save_path,str):
        __save_to_path(save_path)
    else:
        print("Invalid save_path: ",save_path)
        pass 
    pass
if __name__=='__main__':
    pass
