#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import shutil
from pandas import DataFrame
from threading import Thread

from autk.gentk.funcs import f2list,save_df,start_thread_list, get_time_str
from autk.meta.handf.findfile import find_regex

def scanxl(xl_dir=os.path.abspath(os.curdir)):
    from openpyxl import load_workbook
    from pandas import read_excel
    f=find_regex(r'\.xlsx',xl_dir)[0]
    for i in f:
        print('='*5)
        print('bookName:',i.split(os.sep)[-1])
        shtli=load_workbook(i).sheetnames
        print('sheets:',shtli)
        for j in shtli:
            print('\t','sheetName:',j)
            print('\t'*2,'sheetCols:',read_excel(i,sheet_name=j,engine='openpyxl').columns)
        print('-'*5)
        continue
    pass
def tidy_up(
    transform_func, # function to transform jr_str into regex so as to match/search in `search_dir`;
    jrli_path, # the file, each line of whom stores jrs to find;
    search_dir, # 'from' and 'search' directory;
    saveto_dir, # 'to' and save/copy-to directory;
    move=False, # whether to move or copy find-results to `saveto_dir`;
    create_dir=False, # whether to create new directory by the name of `jr_str` beside `saveto_dir` for multi-find-results of files;
    datapath=None, # where to save `jr-location` mapping data;
    preview=True
):
    '''
    Input which jrs to search, by file at `jrli_path`; 
    Tansform jrs into regular expression through `transform_func`;
    Search jrs by regex in `search_dir`;
    Copy/move results into `saveto_dir`;
    Output jr-location mapping data into Excel file at `datapath`;
    -------
    parameters:
        transform_func, # function to transform jr_str into regex;
        jrli_path, # the file which lists which jrs to find;
        search_dir, # 'from' and 'search' directory;
        saveto_dir, # 'to' and save/copy-to directory;
        move=False, # whether to move or copy find-results to `saveto_dir`;
        create_dir=False, # whether to create new directory by the
        name of `jr_str` beside `saveto_dir` for multi-find-results
        of files; directory will be created anyway if more than one
        results found;
        datapath=None, # where to save `jr-location` mapping data;
    returns:
        DataFrame, whose columns are:
        ['target','regex','count','past_path','location']
    '''
    jrli=f2list(jrli_path)
    data_cols=['target','regex','count','past_path','location']
    data=[]
    def single_tidy(jr_str):
        target=jr_str
        regex=transform_func(jr_str)
        from_path_list=find_regex(
            transform_func(jr_str),
            search_dir=search_dir,
            match=True
        )[0] # file results;
        from_dir_list=find_regex(
            transform_func(jr_str),
            search_dir=search_dir,
            match=True
        )[1] # directory results;
        past_path=';'.join(from_path_list)
        past_path=past_path+';'+';'.join(from_dir_list)
        def mvcp_list(resu_list):
            count=len(resu_list)
            def preview_move(from_path,to_path):
                print(jr_str,'-->',from_path.split(os.sep)[-1],',save_to:',to_path)
                pass
            if preview==True:
                process_func=preview_move
            else:
                if move==True:
                    process_func=shutil.move
                else:
                    process_func=shutil.copy
            if len(resu_list)==0:
                #  status='×'
                location=''
                pass
            else:
                #  status='√'
                if len(resu_list)>1 or create_dir==True:
                    # directory will be created anyway 
                    # if more than one results found;
                    save_dir=os.path.abspath(
                        os.path.join(
                            saveto_dir,
                            jr_str+'_rslts' # results for short, as suffix;
                        )
                    )
                    if preview==True:
                        pass
                    else:
                        if os.path.isdir(save_dir):
                            pass
                        else:
                            os.mkdir(save_dir)
                    location=save_dir
                else:
                    # len(resu_list)=1
                    if create_dir==True:
                        save_dir=os.path.abspath(
                            os.path.join(
                                saveto_dir,
                                jr_str+'_rslts'
                            )
                        )
                        if preview==True:
                            pass
                        else:
                            if os.path.isdir(save_dir)==True:
                                pass
                            else:
                                os.mkdir(save_dir)
                    else:
                        save_dir=os.path.abspath(saveto_dir)
                    # get location:
                    location=os.path.join(
                        save_dir,
                        re.sub(
                            r'^.+\.',
                            target+r'.',
                            resu_list[0].split(os.sep)[-1]
                        )
                    )
                thread_list=[]
                for resu_path in resu_list:
                    if len(resu_list)>1:
                        #  print(
                            #  'find multi for',
                            #  jr_str,
                            #  ':',
                            #  len(resu_list)
                        #  )
                        save_path=os.path.join(
                            save_dir,
                            resu_path.split(os.sep)[-1]
                        )
                    else:
                        #  print(
                            #  'find unique for',
                            #  jr_str,
                            #  ':',
                            #  len(resu_list)
                        #  )
                        save_path=location
                    thread_list.append(
                        Thread(
                            target=process_func,
                            args=(resu_path,save_path),
                            name='single_mvcp_'+target.split(os.sep)[-1]
                        )
                    )
                    continue
                start_thread_list(thread_list)
            row_data=[target,regex,count,past_path,location]
            data.append(row_data)
            pass
        # start to process with search-results:
        from_list=from_path_list+from_dir_list
        mvcp_list(from_list)
        return
    # start to tidy up:
    thread_list=[]
    for jr_str in jrli:
        thread_list.append(
            Thread(
                target=single_tidy,
                args=(jr_str,),
                name='copy_jr_'+jr_str
            )
        )
        continue
    start_thread_list(thread_list)
    # get location data:
    data=DataFrame(data,columns=data_cols)
    data.sort_values(
        'target',
        ascending=True,
        inplace=True,
        ignore_index=True
    )
    print(data)
    if datapath is not None and preview==False:
        save_df(data,get_time_str(woc=False),datapath)
    return data
def gendirs(target_dir):
    '''
    This function seems useless now.
    make directories according to the name of files in the current directory.
    去掉所有的中文字符,以英文名字创建文件夹,用以整理资料.
    '''
    for file_name in os.listdir(path=target_dir):
        pure_file_name=re.sub(r'\.pdf$',r'',file_name)
        no_zh_dir_name=re.sub(r'-*[\u4e00-\u9fa5]+.+',r'',pure_file_name)
        print('replace 1: %s'%pure_file_name)
        print('replace 2: %s'%no_zh_dir_name)
        try:
            os.makedirs(no_zh_dir_name)
        except FileExistsError:
            print('%s exists!'%no_zh_dir_name)
            pass
def wcp(
        item,
        frdir,
        todir=os.path.abspath(os.curdir),
        nickName=get_time_str()
):
    '''
    Only files, not directories can be copied.
    frdir is the 'from and search' directory, todir is the 'to' directory.
    Easy copy function using with the find function and then yelling out a big wocao.
    Works well with the find function when you need to copy all the files that you find to a target new directory.
    '''
    import shutil
    cp_files=find_regex(item,frdir)[0]
    file_name='-'.join([nickName,str(frdir.split(os.sep)[-1])])
    destination_dir=os.path.abspath(os.path.join(todir,file_name))
    try:
        shutil.copy(frdir,destination_dir)
    except:
        shutil.copytree(frdir,destination_dir)
    else:
        pass
    finally:
        print('%s is copied.'%frdir)
    return
if __name__=='__main__':
    pass
