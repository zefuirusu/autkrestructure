#!/usr/bin/env python
# coding=utf-8
'''
头一次遇到被审计单位将所有会计主体的余额表都放在一张sheet上的情况，没关系，按主体分开并不难。
'''
from pandas import read_excel
from openpyxl import load_workbook
def separate_tb(tb_df,key_col,savedir='./separate_tb.xlsx'):
    entity_list=list(tb_df[key_col].drop_duplicates())
    def get_entity_tb():
        import re
        for i in entity_list:
            if re.search(re.compile(r'-CAR.*$'),str(i)) !=  None:
            # if 1==1:
                # entity_name=i
                entity_name=re.sub(r'-CAR.*$','',i)
                entity_df=tb_df[tb_df[key_col]==i]
                entity=[entity_name,entity_df]
                yield entity
    from pandas import ExcelWriter
    from openpyxl import Workbook
    wb=Workbook()
    wb.save(savedir)
    wter=ExcelWriter(savedir,book=load_workbook(savedir),engine='openpyxl')
    for i in get_entity_tb():
        i[1].to_excel(wter,sheet_name=i[0])
        wter.save()
        continue
    # wter.save()
    wter.close()
    pass
if __name__=='__main__':
    # tbdir='./2019tb.xlsx'
    # tb=read_excel(tbdir,sheet_name='新的工作表',header=7,engine='openpyxl')
    # separate_tb(tb,'核算账簿名称')
    pass
