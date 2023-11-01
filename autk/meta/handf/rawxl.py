#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
Seems useless, due to the autk.brother.xlsht;
'''

import re
import os
from openpyxl import load_workbook
from xlrd import open_workbook
from pandas import DataFrame
from threading import Thread

from autk.gentk.funcs import start_thread_list

def get_xl_value(fpath,sheet_name,row_index,col_index):
    xl_type=re.sub(
        re.compile(r'^.*\.'),
        '',
        fpath.split(os.sep)[-1]
    )
    if xl_type=='xlsx':
        return load_workbook(fpath)[sheet_name].cell(row_index+1,col_index+1).value
    elif xl_type=='xls':
        return open_workbook(fpath).sheet_by_name(sheet_name).cell(row_index,col_index).value
    else:
        return 'woc'
def draw_xl_value(row_index,col_index,xldir):
    '''
    Get cell value at the same location of all sheets in all workbooks in the directory.
    Note that row_index and col_index start at zero.
    '''
    xl_list=[]
    for f in os.listdir(xldir):
        xl_list.append(
            os.path.join(
                xldir,
                f
            )
        )
        continue
    print(len(xl_list),' files in total.')
    def get_type(fpath):
        return re.sub(
            re.compile(r'^.*\.'),
            '',
            fpath.split(os.sep)[-1]
        )
    def get_value_dict_list(fpath):
        xl_type=get_type(fpath)
        if xl_type=='xls':
            xl=open_workbook(fpath)
            shtli=xl.sheets()
            data=[]
            for sht in shtli:
                data.append(
                    {
                        "book":fpath.split(os.sep)[-1],
                        "sheet":sht.name,
                        "value":xl.sheet_by_name(sht.name).cell(row_index,col_index).value
                    }
                )
                continue
            pass
        elif xl_type=='xlsx':
            xl=load_workbook(fpath)
            shtli=xl.sheetnames()
            data=[]
            for sht in shtli:
                data.append(
                    {
                        "book":fpath.split(os.sep)[-1],
                        "sheet":sht.name,
                        "value":xl.get_sheet_by_name(sht.name).cell(row_index+1,col_index+1).value
                    }
                )
                continue
        else:
            pass
        return data
    def single_parse(collect_list,fpath):
        collect_list.extend(
            get_value_dict_list(fpath)
        )
    thread_list=[]
    data=[]
    for f in xl_list:
        thread_list.append(
            Thread(
                target=single_parse,
                args=(data,f),
                name='get_cell_'+str((row_index,col_index))+'_from_'+f.split(os.sep)[-1]
            )
        )
        #  print(thread_list[-1].name)
        continue
    print(len(thread_list),' threads to start.')
    start_thread_list(thread_list)
    data=DataFrame(data)
    return data
if __name__=='__main__':
    pass
