#!/usr/bin/env python
# coding=utf-8
'''
将分散在各个底稿中的六大往来科目账龄明细拽出来汇总到独立的Excel文件中.
带格式复制哦.
'''
import copy
import os
import re
import threading
from openpyxl import load_workbook,Workbook
from autk import find_regex
# from autk.core.table import ImmortalTable
# datadir='./data'
datadir='./fin沈阳美德因底稿'
outputdir='./output'
tp_li=[
    # '应付账款',
    # '预付款项',
    # '应收账款',
    # '预收款项',
    # '其他应收款',
    # '其他应付款'
]
class TypeWangl:
    def __init__(self,tp):
        self.tp=tp # type, tp for short;
        self._wlist=[]
        self.savepath=os.path.join(outputdir,tp+'.xlsx')
        self.wb=Workbook()
    def save(self):
        self.wb.save(self.savepath)
    def load_bkli(self,wl_list):
        for i in wl_list:
            self._wlist.append(i)
            continue
        pass
    def new_sheet(self,entity_name):
        self.wb.create_sheet(entity_name)
        self.save
        pass
    def copy_sheet(self,source_sheet,target_sheet):
        for i,row in enumerate(source_sheet.iter_rows()):
            for j,cell in enumerate(row):
                target_sheet.cell(row=i+1,column=j+1,value=cell.value)
                if cell.has_style:
                    target_sheet.cell(row=i+1,column=j+1)._style=copy.copy(cell._style)
                    target_sheet.cell(row=i+1,column=j+1).font=copy.copy(cell.font)
                    target_sheet.cell(row=i+1,column=j+1).border=copy.copy(cell.border)
                    target_sheet.cell(row=i+1,column=j+1).fill=copy.copy(cell.fill)
                    target_sheet.cell(row=i+1,column=j+1).number_format=copy.copy(cell.number_format)
                    target_sheet.cell(row=i+1,column=j+1).protection=copy.copy(cell.protection)
                    target_sheet.cell(row=i+1,column=j+1).alignment=copy.copy(cell.alignment)
        self.save()
        pass
    def start(self):
        for w in self._wlist:
            print('copying ...',w.entity_name,' from ',w.path.split(os.sep)[-1])
            self.new_sheet(w.entity_name)
            self.copy_sheet(w.sht,self.wb[w.entity_name])
        pass
    pass
class Wangl:
    def __init__(self,single_path,tp='应付账款'):
        self.path=single_path
        if tp=='预收款项':
            self.sheet_name='预收账款明细表'
        else:
            self.sheet_name=tp+'明细表'
        self.bk=load_workbook(self.path,keep_vba=True)
        self.sht=self.bk[self.sheet_name]
        self.shtli=self.bk.sheetnames
        # self._entity_fake=self.path.split(os.sep)[-1]
        # self.entity_name=re.sub(r'\.xls.?$','',self._entity_fake)
        self.entity_name=self.sht['b2'].value
        pass
    pass
def getwlpth(wl_type_str):
    resu=find_regex(wl_type_str,datadir)[0]
    # print(resu)
    return resu
def join_once(tp_str):
    twl=TypeWangl(tp_str)
    fileli=getwlpth(tp_str)
    for i in fileli:
        w=Wangl(i,tp=tp_str)
        twl._wlist.append(w)
        continue
    twl.start()
def start():
    thread_list=[]
    for i in tp_li:
        t=threading.Thread(target=join_once,args=(i,))
        thread_list.append(t)
        continue
    for t in thread_list:
        t.start()
    for t in thread_list:
        t.join()
    pass
if __name__=='__main__':
    # start()
    pass
